"""Task 2 entry point: compare AI-generated reviews against human reviews.

Steps:
  1. For each paper, match it in gen_review.db and fetch human reviews.
  2. Normalize + align both sides into topic rows (cached per paper).
  3. Score each topic row with an LLM (0/1/2).
  4. Write comparison_output.xlsx.

Usage:
    python scripts/compare_reviews.py

Prerequisites:
  - gen_review.db in project root (or set GEN_REVIEW_DB env var)
  - reviews/ directory populated by task 1 (run_reviews.py)
"""
from __future__ import annotations
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Ensure src/ is on the path when running as a script.
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from reviewer.config import load_config
from reviewer import db as db_mod
from reviewer import normalizer, comparison

_NOT_COVERED = "— not covered —"
_SCORE_FILL = PatternFill("solid", fgColor="D9D9D9")   # gray for AI score
_HEADER_FILL = PatternFill("solid", fgColor="2F5496")  # dark blue header
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _load_ai_review(review_json: Path) -> dict:
    return json.loads(review_json.read_text(encoding="utf-8"))


def _process_paper(
    slug: str,
    review_json: Path,
    conn,
    cfg,
) -> list[dict]:
    ai_review = _load_ai_review(review_json)
    title = ai_review.get("paper", {}).get("title", "")

    paper = db_mod.match_paper(conn, title)
    if not paper:
        print(f"  [WARN] {slug}: no DB match for title: {title!r:.80}")
        return []

    human_reviews = db_mod.get_human_reviews(conn, paper[db_mod._PAPER_ID_COL])
    if not human_reviews:
        print(f"  [WARN] {slug}: no human reviews found in DB")
        return []

    # Use cached normalized topics if available.
    cache_path = review_json.parent / "normalized_topics.json"
    if cache_path.exists():
        print(f"  {slug}: using cached normalized_topics.json")
        topic_rows = json.loads(cache_path.read_text(encoding="utf-8"))
    else:
        print(f"  {slug}: normalizing {len(human_reviews)} human review(s)…")
        topic_rows = normalizer.normalize_and_align(
            human_reviews, ai_review, cfg.model, db_mod.review_prose
        )
        cache_path.write_text(
            json.dumps(topic_rows, indent=2, ensure_ascii=False), encoding="utf-8"
        )

    print(f"  {slug}: scoring {len(topic_rows)} topic(s)…")
    scored = comparison.score_topic_rows(topic_rows, cfg.model)

    rows = []
    for row in scored:
        rows.append(
            {
                "Paper ID": slug,
                "Paper Title": title,
                "Topic": row.get("topic", ""),
                "Human Reviews (All Points)": row.get("human_all") or _NOT_COVERED,
                "AI View": row.get("ai_view") or _NOT_COVERED,
                "AI Score": row.get("suggested_score", 0),
                "AI Reasoning": row.get("reasoning", ""),
                "User Label": "",
                "Notes": "",
            }
        )
    return rows


def _write_excel(all_rows: list[dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review Comparison"

    headers = [
        "Paper ID",
        "Paper Title",
        "Topic",
        "Human Reviews (All Points)",
        "AI View",
        "AI Score",
        "AI Reasoning",
        "User Label",
        "Notes",
    ]

    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for row_idx, row in enumerate(all_rows, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if header == "AI Score":
                cell.fill = _SCORE_FILL

    # Dropdown validation on "User Label" column (col 8)
    user_label_col = headers.index("User Label") + 1
    dv = DataValidation(
        type="list",
        formula1='"0,1,2"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = (
        f"{get_column_letter(user_label_col)}2:"
        f"{get_column_letter(user_label_col)}{len(all_rows) + 1}"
    )
    ws.add_data_validation(dv)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Column widths (approximate)
    widths = {
        "Paper ID": 10,
        "Paper Title": 40,
        "Topic": 22,
        "Human Reviews (All Points)": 55,
        "AI View": 55,
        "AI Score": 10,
        "AI Reasoning": 50,
        "User Label": 12,
        "Notes": 30,
    }
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 20)

    # Row heights for readability
    for row_idx in range(2, len(all_rows) + 2):
        ws.row_dimensions[row_idx].height = 80

    wb.save(output_path)


def main() -> None:
    cfg = load_config()

    if not cfg.gen_review_db.exists():
        sys.exit(
            f"Error: gen_review.db not found at '{cfg.gen_review_db}'.\n"
            "Download the Gen-Review dataset (DOI: 10.7910/DVN/PYDPEZ) and "
            "place gen_review.db in the project root, or set GEN_REVIEW_DB."
        )

    conn = db_mod.connect(cfg.gen_review_db)
    all_rows: list[dict] = []

    review_dirs = sorted(
        p for p in cfg.reviews_dir.iterdir()
        if p.is_dir() and (p / "review.json").exists()
    )

    if not review_dirs:
        sys.exit(f"No review.json files found under '{cfg.reviews_dir}'. Run task 1 first.")

    print(f"Processing {len(review_dirs)} paper(s) from '{cfg.reviews_dir}'…\n")

    for review_dir in review_dirs:
        slug = review_dir.name
        print(f"[{slug}]")
        rows = _process_paper(slug, review_dir / "review.json", conn, cfg)
        all_rows.extend(rows)
        print()

    conn.close()

    if not all_rows:
        sys.exit("No rows produced — check warnings above.")

    output = Path("comparison_output.xlsx")
    _write_excel(all_rows, output)
    print(f"Wrote {len(all_rows)} rows to {output}")


if __name__ == "__main__":
    main()
